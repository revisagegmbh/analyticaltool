"""
Data Manager Module - JSON Persistence & DataFrame Management
Handles all data operations, storage, and analytics calculations
"""

import json
import os
from datetime import datetime
from pathlib import Path
from typing import Optional, Dict, List, Tuple, Any
import pandas as pd
import numpy as np


class DataManager:
    """Manages expense data with JSON persistence."""
    
    # Default data directory (relative to script location for Desktop App compatibility)
    DEFAULT_DATA_DIR = Path(__file__).parent / "data"
    DEFAULT_DATA_FILE = "expenses.json"
    
    def __init__(self, data_dir: Optional[Path] = None):
        """Initialize the data manager.
        
        Args:
            data_dir: Custom data directory path. Uses default if None.
        """
        self.data_dir = Path(data_dir) if data_dir else self.DEFAULT_DATA_DIR
        self.data_file = self.data_dir / self.DEFAULT_DATA_FILE
        
        # Ensure data directory exists
        self.data_dir.mkdir(parents=True, exist_ok=True)
        
        # Initialize DataFrame with extended columns
        self.expenses_df = pd.DataFrame(columns=[
            'Date', 'Amount', 'Description', 'Source', 
            'Category', 'Vendor', 'Currency', 'ID'
        ])
        
        # Load existing data
        self.load_data()
    
    def load_data(self) -> bool:
        """Load expenses from JSON file.
        
        Returns:
            True if data was loaded successfully, False otherwise.
        """
        if not self.data_file.exists():
            return False
        
        try:
            with open(self.data_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            if data.get('expenses'):
                self.expenses_df = pd.DataFrame(data['expenses'])
                self.expenses_df['Date'] = pd.to_datetime(self.expenses_df['Date'])
                self.expenses_df = self.expenses_df.sort_values('Date')
                return True
        except (json.JSONDecodeError, KeyError, Exception) as e:
            print(f"Error loading data: {e}")
        
        return False
    
    def save_data(self) -> bool:
        """Save expenses to JSON file.
        
        Returns:
            True if data was saved successfully, False otherwise.
        """
        try:
            # Prepare data for JSON serialization
            df_copy = self.expenses_df.copy()
            df_copy['Date'] = df_copy['Date'].dt.strftime('%Y-%m-%d')
            
            data = {
                'version': '1.0',
                'last_updated': datetime.now().isoformat(),
                'expenses': df_copy.to_dict('records')
            }
            
            with open(self.data_file, 'w', encoding='utf-8') as f:
                json.dump(data, f, indent=2, ensure_ascii=False)
            
            return True
        except Exception as e:
            print(f"Error saving data: {e}")
            return False
    
    def add_expenses(self, expenses: List[Dict]) -> int:
        """Add new expenses to the DataFrame.
        
        Args:
            expenses: List of expense dictionaries.
            
        Returns:
            Number of expenses added.
        """
        if not expenses:
            return 0
        
        # Generate unique IDs for new expenses
        for expense in expenses:
            if 'ID' not in expense or not expense['ID']:
                expense['ID'] = self._generate_id()
            if 'Category' not in expense:
                expense['Category'] = 'Uncategorized'
            if 'Vendor' not in expense:
                expense['Vendor'] = ''
            if 'Currency' not in expense:
                expense['Currency'] = 'EUR'
        
        new_df = pd.DataFrame(expenses)
        new_df['Date'] = pd.to_datetime(new_df['Date'])
        
        self.expenses_df = pd.concat([self.expenses_df, new_df], ignore_index=True)
        self.expenses_df = self.expenses_df.sort_values('Date')
        
        # Auto-save
        self.save_data()
        
        return len(expenses)
    
    def _generate_id(self) -> str:
        """Generate a unique expense ID."""
        import uuid
        return str(uuid.uuid4())[:8]
    
    def clear_data(self):
        """Clear all expense data."""
        self.expenses_df = pd.DataFrame(columns=[
            'Date', 'Amount', 'Description', 'Source', 
            'Category', 'Vendor', 'Currency', 'ID'
        ])
        self.save_data()
    
    def get_date_range(self) -> Tuple[Optional[datetime], Optional[datetime]]:
        """Get the date range of all expenses.
        
        Returns:
            Tuple of (min_date, max_date) or (None, None) if no data.
        """
        if self.expenses_df.empty:
            return None, None
        
        return self.expenses_df['Date'].min(), self.expenses_df['Date'].max()
    
    def filter_by_date_range(self, start_date: datetime, end_date: datetime) -> pd.DataFrame:
        """Filter expenses by date range.
        
        Args:
            start_date: Start date (inclusive).
            end_date: End date (inclusive).
            
        Returns:
            Filtered DataFrame.
        """
        if self.expenses_df.empty:
            return self.expenses_df.copy()
        
        mask = (self.expenses_df['Date'] >= pd.Timestamp(start_date)) & \
               (self.expenses_df['Date'] <= pd.Timestamp(end_date))
        return self.expenses_df[mask].copy()
    
    # =========================================================================
    # KPI Calculations
    # =========================================================================
    
    def calculate_kpis(self, start_date: Optional[datetime] = None, 
                       end_date: Optional[datetime] = None) -> Dict[str, Any]:
        """Calculate key performance indicators for the selected period.
        
        Args:
            start_date: Start date for filtering.
            end_date: End date for filtering.
            
        Returns:
            Dictionary with KPI values.
        """
        # Get filtered data
        if start_date and end_date:
            df = self.filter_by_date_range(start_date, end_date)
        else:
            df = self.expenses_df.copy()
        
        if df.empty:
            return {
                'total_expenses': 0,
                'avg_monthly': 0,
                'highest_month': ('N/A', 0),
                'lowest_month': ('N/A', 0),
                'yoy_growth': None,
                'mom_change': None,
                'invoice_count': 0,
            }
        
        # Total expenses
        total_expenses = df['Amount'].sum()
        
        # Monthly aggregation
        monthly = df.copy()
        monthly['YearMonth'] = monthly['Date'].dt.to_period('M')
        monthly_totals = monthly.groupby('YearMonth')['Amount'].sum()
        
        # Average monthly expense
        avg_monthly = monthly_totals.mean() if len(monthly_totals) > 0 else 0
        
        # Highest and lowest months
        if len(monthly_totals) > 0:
            highest_idx = monthly_totals.idxmax()
            lowest_idx = monthly_totals.idxmin()
            highest_month = (str(highest_idx), monthly_totals[highest_idx])
            lowest_month = (str(lowest_idx), monthly_totals[lowest_idx])
        else:
            highest_month = ('N/A', 0)
            lowest_month = ('N/A', 0)
        
        # Year-over-Year growth
        yoy_growth = self._calculate_yoy_growth(df)
        
        # Month-over-Month change
        mom_change = self._calculate_mom_change(df)
        
        return {
            'total_expenses': total_expenses,
            'avg_monthly': avg_monthly,
            'highest_month': highest_month,
            'lowest_month': lowest_month,
            'yoy_growth': yoy_growth,
            'mom_change': mom_change,
            'invoice_count': len(df),
        }
    
    def _calculate_yoy_growth(self, df: pd.DataFrame) -> Optional[float]:
        """Calculate year-over-year growth percentage.
        
        Returns:
            YoY growth as percentage, or None if not calculable.
        """
        if df.empty:
            return None
        
        yearly = df.copy()
        yearly['Year'] = yearly['Date'].dt.year
        yearly_totals = yearly.groupby('Year')['Amount'].sum()
        
        if len(yearly_totals) < 2:
            return None
        
        years = sorted(yearly_totals.index)
        current_year = years[-1]
        previous_year = years[-2]
        
        current = yearly_totals[current_year]
        previous = yearly_totals[previous_year]
        
        if previous == 0:
            return None
        
        return ((current - previous) / previous) * 100
    
    def _calculate_mom_change(self, df: pd.DataFrame) -> Optional[Dict]:
        """Calculate month-over-month change.
        
        Returns:
            Dict with 'percentage' and 'absolute' change, or None.
        """
        if df.empty:
            return None
        
        monthly = df.copy()
        monthly['YearMonth'] = monthly['Date'].dt.to_period('M')
        monthly_totals = monthly.groupby('YearMonth')['Amount'].sum()
        
        if len(monthly_totals) < 2:
            return None
        
        months = sorted(monthly_totals.index)
        current_month = months[-1]
        previous_month = months[-2]
        
        current = monthly_totals[current_month]
        previous = monthly_totals[previous_month]
        
        absolute_change = current - previous
        
        if previous == 0:
            percentage_change = 100 if current > 0 else 0
        else:
            percentage_change = ((current - previous) / previous) * 100
        
        return {
            'percentage': percentage_change,
            'absolute': absolute_change,
            'current_month': str(current_month),
            'previous_month': str(previous_month),
        }
    
    # =========================================================================
    # Aggregation Methods
    # =========================================================================
    
    def get_monthly_totals(self, start_date: Optional[datetime] = None,
                           end_date: Optional[datetime] = None) -> pd.DataFrame:
        """Get monthly expense totals.
        
        Returns:
            DataFrame with YearMonth and Amount columns.
        """
        if start_date and end_date:
            df = self.filter_by_date_range(start_date, end_date)
        else:
            df = self.expenses_df.copy()
        
        if df.empty:
            return pd.DataFrame(columns=['YearMonth', 'Amount'])
        
        monthly = df.copy()
        monthly['YearMonth'] = monthly['Date'].dt.to_period('M')
        grouped = monthly.groupby('YearMonth')['Amount'].sum().reset_index()
        grouped['YearMonth'] = grouped['YearMonth'].astype(str)
        
        return grouped
    
    def get_yearly_totals(self, start_date: Optional[datetime] = None,
                          end_date: Optional[datetime] = None) -> pd.DataFrame:
        """Get yearly expense totals.
        
        Returns:
            DataFrame with Year and Amount columns.
        """
        if start_date and end_date:
            df = self.filter_by_date_range(start_date, end_date)
        else:
            df = self.expenses_df.copy()
        
        if df.empty:
            return pd.DataFrame(columns=['Year', 'Amount'])
        
        yearly = df.copy()
        yearly['Year'] = yearly['Date'].dt.year
        grouped = yearly.groupby('Year')['Amount'].sum().reset_index()
        
        return grouped
    
    def get_monthly_comparison(self) -> pd.DataFrame:
        """Get month-over-month comparison data.
        
        Returns:
            DataFrame with monthly data and percentage changes.
        """
        if self.expenses_df.empty:
            return pd.DataFrame()
        
        monthly = self.expenses_df.copy()
        monthly['YearMonth'] = monthly['Date'].dt.to_period('M')
        grouped = monthly.groupby('YearMonth')['Amount'].sum().reset_index()
        
        # Calculate percentage change
        grouped['PrevAmount'] = grouped['Amount'].shift(1)
        grouped['Change'] = grouped['Amount'] - grouped['PrevAmount']
        grouped['ChangePercent'] = (grouped['Change'] / grouped['PrevAmount']) * 100
        grouped['ChangePercent'] = grouped['ChangePercent'].fillna(0)
        
        grouped['YearMonth'] = grouped['YearMonth'].astype(str)
        
        return grouped
    
    def get_trend_data(self, window: int = 3) -> Tuple[List, List]:
        """Calculate moving average for trend line.
        
        Args:
            window: Moving average window size.
            
        Returns:
            Tuple of (x_values, trend_values).
        """
        monthly = self.get_monthly_totals()
        
        if len(monthly) < window:
            return [], []
        
        trend = monthly['Amount'].rolling(window=window, center=True).mean()
        
        return list(range(len(trend))), trend.tolist()
    
    # =========================================================================
    # Date Range Presets
    # =========================================================================
    
    @staticmethod
    def get_date_range_presets() -> Dict[str, Tuple[datetime, datetime]]:
        """Get predefined date range options.
        
        Returns:
            Dictionary of preset name to (start_date, end_date) tuples.
        """
        today = datetime.now()
        
        # Calculate quarters
        current_quarter = (today.month - 1) // 3 + 1
        q1_start = datetime(today.year, 1, 1)
        q1_end = datetime(today.year, 3, 31)
        q2_start = datetime(today.year, 4, 1)
        q2_end = datetime(today.year, 6, 30)
        q3_start = datetime(today.year, 7, 1)
        q3_end = datetime(today.year, 9, 30)
        q4_start = datetime(today.year, 10, 1)
        q4_end = datetime(today.year, 12, 31)
        
        # Previous year quarters
        prev_year = today.year - 1
        
        presets = {
            'Gesamter Zeitraum': (datetime(2000, 1, 1), today),
            'Dieses Jahr': (datetime(today.year, 1, 1), today),
            'Letztes Jahr': (datetime(prev_year, 1, 1), datetime(prev_year, 12, 31)),
            'Letzte 12 Monate': (datetime(today.year - 1, today.month, 1), today),
            'Letzte 6 Monate': (datetime(today.year if today.month > 6 else today.year - 1,
                                         (today.month - 6) % 12 or 12, 1), today),
            'Letzte 3 Monate': (datetime(today.year if today.month > 3 else today.year - 1,
                                         (today.month - 3) % 12 or 12, 1), today),
            'Dieser Monat': (datetime(today.year, today.month, 1), today),
            f'Q1 {today.year}': (q1_start, q1_end),
            f'Q2 {today.year}': (q2_start, q2_end),
            f'Q3 {today.year}': (q3_start, q3_end),
            f'Q4 {today.year}': (q4_start, q4_end),
            f'Q1 {prev_year}': (datetime(prev_year, 1, 1), datetime(prev_year, 3, 31)),
            f'Q2 {prev_year}': (datetime(prev_year, 4, 1), datetime(prev_year, 6, 30)),
            f'Q3 {prev_year}': (datetime(prev_year, 7, 1), datetime(prev_year, 9, 30)),
            f'Q4 {prev_year}': (datetime(prev_year, 10, 1), datetime(prev_year, 12, 31)),
            'Benutzerdefiniert': (None, None),  # Placeholder for custom range
        }
        
        return presets
    
    # =========================================================================
    # Category Management (Phase 3)
    # =========================================================================
    
    # Predefined categories for invoices
    CATEGORIES = [
        'Uncategorized',
        'Dienstleistungen',
        'Material',
        'Transport',
        'Miete',
        'Versicherung',
        'Marketing',
        'IT & Software',
        'Bürobedarf',
        'Reisekosten',
        'Beratung',
        'Wartung',
        'Sonstiges'
    ]
    
    def update_expense_category(self, expense_id: str, category: str) -> bool:
        """Update the category of an expense.
        
        Args:
            expense_id: The ID of the expense to update.
            category: The new category.
            
        Returns:
            True if successful, False otherwise.
        """
        if self.expenses_df.empty:
            return False
        
        mask = self.expenses_df['ID'] == expense_id
        if mask.any():
            self.expenses_df.loc[mask, 'Category'] = category
            self.save_data()
            return True
        return False
    
    def update_expense_vendor(self, expense_id: str, vendor: str) -> bool:
        """Update the vendor of an expense.
        
        Args:
            expense_id: The ID of the expense to update.
            vendor: The new vendor name.
            
        Returns:
            True if successful, False otherwise.
        """
        if self.expenses_df.empty:
            return False
        
        mask = self.expenses_df['ID'] == expense_id
        if mask.any():
            self.expenses_df.loc[mask, 'Vendor'] = vendor
            self.save_data()
            return True
        return False
    
    def get_categories_summary(self) -> pd.DataFrame:
        """Get summary of expenses by category.
        
        Returns:
            DataFrame with Category, Count, and Total columns.
        """
        if self.expenses_df.empty:
            return pd.DataFrame(columns=['Category', 'Count', 'Total'])
        
        summary = self.expenses_df.groupby('Category').agg(
            Count=('Amount', 'count'),
            Total=('Amount', 'sum')
        ).reset_index()
        
        return summary.sort_values('Total', ascending=False)
    
    # =========================================================================
    # Search and Filter (Phase 3)
    # =========================================================================
    
    def search_expenses(self, query: str, 
                       start_date: Optional[datetime] = None,
                       end_date: Optional[datetime] = None,
                       category: Optional[str] = None,
                       min_amount: Optional[float] = None,
                       max_amount: Optional[float] = None) -> pd.DataFrame:
        """Search and filter expenses.
        
        Args:
            query: Search text (searches in Description, Source, Vendor).
            start_date: Filter by start date.
            end_date: Filter by end date.
            category: Filter by category.
            min_amount: Minimum amount filter.
            max_amount: Maximum amount filter.
            
        Returns:
            Filtered DataFrame.
        """
        if self.expenses_df.empty:
            return self.expenses_df.copy()
        
        df = self.expenses_df.copy()
        
        # Date filter
        if start_date:
            df = df[df['Date'] >= pd.Timestamp(start_date)]
        if end_date:
            df = df[df['Date'] <= pd.Timestamp(end_date)]
        
        # Category filter
        if category and category != 'Alle Kategorien':
            df = df[df['Category'] == category]
        
        # Amount filter
        if min_amount is not None:
            df = df[df['Amount'] >= min_amount]
        if max_amount is not None:
            df = df[df['Amount'] <= max_amount]
        
        # Text search
        if query and query.strip():
            query_lower = query.lower().strip()
            mask = (
                df['Description'].str.lower().str.contains(query_lower, na=False) |
                df['Source'].str.lower().str.contains(query_lower, na=False) |
                df['Vendor'].str.lower().str.contains(query_lower, na=False)
            )
            df = df[mask]
        
        return df
    
    # =========================================================================
    # Export Functions (Phase 3)
    # =========================================================================
    
    def export_to_csv(self, filepath: str, 
                      start_date: Optional[datetime] = None,
                      end_date: Optional[datetime] = None) -> bool:
        """Export expenses to CSV file.
        
        Args:
            filepath: Path to save the CSV file.
            start_date: Optional start date filter.
            end_date: Optional end date filter.
            
        Returns:
            True if successful, False otherwise.
        """
        try:
            if start_date and end_date:
                df = self.filter_by_date_range(start_date, end_date)
            else:
                df = self.expenses_df.copy()
            
            # Format for export
            export_df = df.copy()
            export_df['Date'] = export_df['Date'].dt.strftime('%d.%m.%Y')
            export_df['Amount'] = export_df['Amount'].apply(lambda x: f"{x:.2f}".replace('.', ','))
            
            # Reorder columns
            columns = ['Date', 'Amount', 'Description', 'Category', 'Vendor', 'Source', 'Currency']
            export_df = export_df[[c for c in columns if c in export_df.columns]]
            
            # German column names
            column_names = {
                'Date': 'Datum',
                'Amount': 'Betrag',
                'Description': 'Beschreibung',
                'Category': 'Kategorie',
                'Vendor': 'Lieferant',
                'Source': 'Quelldatei',
                'Currency': 'Währung'
            }
            export_df = export_df.rename(columns=column_names)
            
            export_df.to_csv(filepath, index=False, encoding='utf-8-sig', sep=';')
            return True
        except Exception as e:
            print(f"CSV Export Error: {e}")
            return False
    
    def export_to_excel(self, filepath: str,
                        start_date: Optional[datetime] = None,
                        end_date: Optional[datetime] = None) -> bool:
        """Export expenses to Excel file with formatting.
        
        Args:
            filepath: Path to save the Excel file.
            start_date: Optional start date filter.
            end_date: Optional end date filter.
            
        Returns:
            True if successful, False otherwise.
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils.dataframe import dataframe_to_rows
            
            if start_date and end_date:
                df = self.filter_by_date_range(start_date, end_date)
            else:
                df = self.expenses_df.copy()
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Einnahmen"
            
            # Styles
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(start_color="1E88E5", end_color="1E88E5", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            currency_alignment = Alignment(horizontal="right")
            date_alignment = Alignment(horizontal="center")
            
            thin_border = Border(
                left=Side(style='thin', color='CCCCCC'),
                right=Side(style='thin', color='CCCCCC'),
                top=Side(style='thin', color='CCCCCC'),
                bottom=Side(style='thin', color='CCCCCC')
            )
            
            # Headers
            headers = ['Datum', 'Betrag (€)', 'Beschreibung', 'Kategorie', 'Lieferant', 'Quelldatei']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # Data rows
            for row_idx, (_, row) in enumerate(df.iterrows(), 2):
                # Date
                cell = ws.cell(row=row_idx, column=1, value=row['Date'].strftime('%d.%m.%Y'))
                cell.alignment = date_alignment
                cell.border = thin_border
                
                # Amount
                cell = ws.cell(row=row_idx, column=2, value=row['Amount'])
                cell.number_format = '#,##0.00 €'
                cell.alignment = currency_alignment
                cell.border = thin_border
                
                # Description
                cell = ws.cell(row=row_idx, column=3, value=row['Description'])
                cell.border = thin_border
                
                # Category
                cell = ws.cell(row=row_idx, column=4, value=row.get('Category', 'Uncategorized'))
                cell.border = thin_border
                
                # Vendor
                cell = ws.cell(row=row_idx, column=5, value=row.get('Vendor', ''))
                cell.border = thin_border
                
                # Source
                cell = ws.cell(row=row_idx, column=6, value=row['Source'])
                cell.border = thin_border
            
            # Summary row
            summary_row = len(df) + 3
            ws.cell(row=summary_row, column=1, value="SUMME:").font = Font(bold=True)
            sum_cell = ws.cell(row=summary_row, column=2, value=f"=SUM(B2:B{len(df)+1})")
            sum_cell.number_format = '#,##0.00 €'
            sum_cell.font = Font(bold=True)
            
            # Column widths
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 14
            ws.column_dimensions['C'].width = 30
            ws.column_dimensions['D'].width = 18
            ws.column_dimensions['E'].width = 25
            ws.column_dimensions['F'].width = 25
            
            wb.save(filepath)
            return True
        except ImportError:
            print("openpyxl nicht installiert. Bitte 'pip install openpyxl' ausführen.")
            return False
        except Exception as e:
            print(f"Excel Export Error: {e}")
            return False
    
    def export_to_pdf(self, filepath: str,
                      start_date: Optional[datetime] = None,
                      end_date: Optional[datetime] = None) -> bool:
        """Export expenses to PDF report.
        
        Args:
            filepath: Path to save the PDF file.
            start_date: Optional start date filter.
            end_date: Optional end date filter.
            
        Returns:
            True if successful, False otherwise.
        """
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import cm
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            
            if start_date and end_date:
                df = self.filter_by_date_range(start_date, end_date)
            else:
                df = self.expenses_df.copy()
            
            doc = SimpleDocTemplate(filepath, pagesize=A4, 
                                   leftMargin=1.5*cm, rightMargin=1.5*cm,
                                   topMargin=2*cm, bottomMargin=2*cm)
            
            elements = []
            styles = getSampleStyleSheet()
            
            # Title
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                spaceAfter=20,
                textColor=colors.HexColor('#1E88E5')
            )
            elements.append(Paragraph("Einnahmen-Report", title_style))
            
            # Date range info
            if start_date and end_date:
                date_info = f"Zeitraum: {start_date.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')}"
            else:
                date_info = f"Erstellt am: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
            elements.append(Paragraph(date_info, styles['Normal']))
            elements.append(Spacer(1, 20))
            
            # Summary
            total = df['Amount'].sum()
            count = len(df)
            summary_text = f"<b>Anzahl Rechnungen:</b> {count}<br/><b>Gesamtsumme:</b> €{total:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
            elements.append(Paragraph(summary_text, styles['Normal']))
            elements.append(Spacer(1, 20))
            
            # Table
            if not df.empty:
                table_data = [['Datum', 'Betrag', 'Beschreibung', 'Kategorie']]
                
                for _, row in df.iterrows():
                    table_data.append([
                        row['Date'].strftime('%d.%m.%Y'),
                        f"€{row['Amount']:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'),
                        row['Description'][:30] + ('...' if len(row['Description']) > 30 else ''),
                        row.get('Category', 'Uncategorized')
                    ])
                
                table = Table(table_data, colWidths=[2.5*cm, 3*cm, 7*cm, 4*cm])
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E88E5')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
                    ('ALIGN', (2, 1), (2, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('FONTSIZE', (0, 1), (-1, -1), 9),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F5F5F5')),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')]),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ]))
                elements.append(table)
            
            doc.build(elements)
            return True
        except ImportError:
            print("reportlab nicht installiert. Bitte 'pip install reportlab' ausführen.")
            return False
        except Exception as e:
            print(f"PDF Export Error: {e}")
            return False

